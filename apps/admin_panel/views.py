# apps/admin_panel/views.py
from datetime import datetime, date, timedelta
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import io
import csv
from apps.core.models import Configuracion, Feriado, Menu
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views import View
from django.db import transaction
from django.core.paginator import Paginator

from apps.accounts.models import CustomUser
from apps.core.models import Precio
from apps.comedor.models import Transaccion, LogCarga, Compra
from apps.pagos.models import CargaVirtual
from .mixins import CajeroRequiredMixin, AdministradorRequiredMixin, RepartidorRequiredMixin


class IndexView(CajeroRequiredMixin, View):
    """
    Panel principal del vendedor.
    Equivale a index() de Vendedor.php en CI3.
    """
    template_name = 'admin_panel/index.html'

    def get(self, request):
        usuario = None
        documento = request.GET.get('documento')

        if documento:
            try:
                usuario = CustomUser.objects.get(documento=int(documento))
            except (CustomUser.DoesNotExist, ValueError):
                messages.error(
                    request, 'No existe un usuario con ese documento.')

        context = {
            'titulo': 'Carga de Saldo',
            'usuario': usuario,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        documento = request.POST.get('numeroDni')
        try:
            usuario = CustomUser.objects.get(documento=int(documento))
            return render(request, self.template_name, {
                'titulo': 'Carga de Saldo',
                'usuario': usuario,
            })
        except (CustomUser.DoesNotExist, ValueError):
            messages.error(request, 'No existe un usuario con ese documento.')
            return redirect('admin_panel:index')


class CargarSaldoView(CajeroRequiredMixin, View):
    """
    Equivale a cargarSaldo() de Vendedor.php en CI3.
    """

    def post(self, request):
        documento = request.POST.get('dni')
        monto = request.POST.get('carga')
        metodo = request.POST.get('metodo_carga')

        if not documento or not monto or not metodo:
            messages.error(request, 'Completá todos los campos.')
            return redirect('admin_panel:index')

        try:
            monto = float(monto)
            if monto == 0:
                messages.error(request, 'El monto debe ser distinto de 0.')
                return redirect('admin_panel:index')
        except ValueError:
            messages.error(request, 'El monto debe ser numérico.')
            return redirect('admin_panel:index')

        try:
            usuario = CustomUser.objects.get(documento=int(documento))
        except CustomUser.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('admin_panel:index')

        with transaction.atomic():
            nuevo_saldo = float(usuario.saldo) + monto
            tipo = 'Carga de Saldo' if monto > 0 else 'Devolucion de Saldo'

            transaccion = Transaccion.objects.create(
                usuario=usuario,
                transaccion=tipo,
                monto=monto,
                saldo=nuevo_saldo,
            )

            LogCarga.objects.create(
                usuario=usuario,
                vendedor=request.user,
                monto=monto,
                formato=metodo,
                transaccion=transaccion,
            )

            usuario.saldo = nuevo_saldo
            usuario.save(update_fields=['saldo'])

        # Enviar email
        self._enviar_email_carga(usuario, transaccion,
                                 metodo, monto, nuevo_saldo)

        messages.success(
            request,
            f'Se cargaron ${monto} a {usuario.nombre_completo}. '
            f'Nuevo saldo: ${nuevo_saldo}'
        )
        return redirect('admin_panel:index')

    def _enviar_email_carga(self, usuario, transaccion, metodo, monto, saldo):
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        from django.conf import settings

        try:
            context = {
                'transaccion': transaccion.id,
                'usuario': usuario,
                'monto': monto,
                'saldo': saldo,
                'metodo': metodo,
            }
            mensaje = render_to_string('emails/carga_saldo.html', context)
            send_mail(
                subject='Carga de Saldo - Comedor UTN FRLP',
                message='',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                html_message=mensaje,
                fail_silently=True,
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error email carga: {e}")


class CrearUsuarioView(CajeroRequiredMixin, View):
    """
    Equivale a createUser() de Vendedor.php en CI3.
    """
    template_name = 'admin_panel/crear_usuario.html'

    def get(self, request):
        precios = Precio.objects.all()
        return render(request, self.template_name, {
            'titulo': 'Nuevo Usuario',
            'precios': precios,
        })

    def post(self, request):
        datos = {
            'documento': request.POST.get('dni'),
            'legajo': request.POST.get('legajo'),
            'first_name': request.POST.get('nombre', '').title(),
            'last_name': request.POST.get('apellido', '').title(),
            'email': request.POST.get('email', '').lower(),
            'tipo': request.POST.get('claustro'),
            'especialidad': request.POST.get('especialidad') or None,
            'es_becado': request.POST.get('beca') == 'Si',
            'saldo': float(request.POST.get('saldo', 0)),
        }

        # Validaciones básicas
        errores = []
        if CustomUser.objects.filter(documento=datos['documento']).exists():
            errores.append('Ese documento ya está registrado.')
        if CustomUser.objects.filter(legajo=datos['legajo']).exists():
            errores.append('Ese legajo ya está registrado.')
        if CustomUser.objects.filter(email=datos['email']).exists():
            errores.append('Ese email ya está registrado.')

        if errores:
            for e in errores:
                messages.error(request, e)
            return render(request, self.template_name, {
                'titulo': 'Nuevo Usuario',
                'precios': Precio.objects.all(),
            })

        # Generar password aleatorio
        import random
        import string
        letras = ''.join(random.choices(string.ascii_lowercase, k=3))
        numeros = ''.join(random.choices(string.digits, k=3))
        password = f"{letras}{numeros}"

        with transaction.atomic():
            usuario = CustomUser.objects.create_user(
                username=str(datos['documento']),
                documento=datos['documento'],
                password=password,
                **{k: v for k, v in datos.items()
                   if k not in ['documento']},
            )

            # Log de alta
            from apps.comedor.models import LogCarga
            if datos['saldo'] > 0:
                transaccion = Transaccion.objects.create(
                    usuario=usuario,
                    transaccion='Carga de Saldo',
                    monto=datos['saldo'],
                    saldo=datos['saldo'],
                )
                LogCarga.objects.create(
                    usuario=usuario,
                    vendedor=request.user,
                    monto=datos['saldo'],
                    formato='Efectivo',
                    transaccion=transaccion,
                )

        # Email bienvenida
        self._enviar_email_bienvenida(usuario, password)

        messages.success(
            request, f'Usuario {usuario.nombre_completo} creado correctamente.')
        return redirect('admin_panel:index')

    def _enviar_email_bienvenida(self, usuario, password):
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        from django.conf import settings

        try:
            context = {'usuario': usuario, 'password': password}
            mensaje = render_to_string('emails/nuevo_usuario.html', context)
            send_mail(
                subject='Bienvenido al Comedor UTN FRLP',
                message='',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                html_message=mensaje,
                fail_silently=True,
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error email bienvenida: {e}")


class ModificarUsuarioView(CajeroRequiredMixin, View):
    """
    Equivale a updateUser() de Vendedor.php en CI3.
    """
    template_name = 'admin_panel/modificar_usuario.html'


    def get(self, request, pk):
        usuario = get_object_or_404(CustomUser, pk=pk)
        return render(request, self.template_name, {
            'titulo': f'Modificar {usuario.nombre_completo}',
            'usuario': usuario,
            'especialidad_choices': CustomUser.ESPECIALIDAD_CHOICES,
        })
    
    def post(self, request, pk):
        usuario = get_object_or_404(CustomUser, pk=pk)

        usuario.first_name = request.POST.get('nombre', '').title()
        usuario.last_name = request.POST.get('apellido', '').title()
        usuario.email = request.POST.get('email', '').lower()
        usuario.tipo = request.POST.get('claustro')
        usuario.especialidad = request.POST.get('especialidad') or None
        usuario.es_becado = request.POST.get('beca') == 'Si'
        usuario.legajo = request.POST.get('legajo')
        usuario.documento = request.POST.get('documento')
        usuario.save()

        messages.success(request, 'Usuario actualizado correctamente.')
        return redirect('admin_panel:index')


class HistorialCargasView(CajeroRequiredMixin, View):
    """
    Equivale a historialCargas() de Vendedor.php en CI3.
    """
    template_name = 'admin_panel/historial_cargas.html'

    def get(self, request):
        cargas = LogCarga.objects.filter(
            vendedor=request.user
        ).select_related('usuario').order_by('-fecha', '-hora')[:20]

        return render(request, self.template_name, {
            'titulo': 'Historial de Cargas',
            'cargas': cargas,
        })


class VerComprasUsuarioView(AdministradorRequiredMixin, View):
    """
    Equivale a ver_compras_userid() de Administrador.php en CI3.
    """
    template_name = 'admin_panel/ver_compras.html'

    def get(self, request, pk):
        usuario = get_object_or_404(CustomUser, pk=pk)
        compras_qs = Compra.objects.filter(
            usuario=usuario
        ).select_related('transaccion').order_by('-dia_comprado')

        paginator = Paginator(compras_qs, 10)
        page = request.GET.get('page', 1)
        compras = paginator.get_page(page)

        return render(request, self.template_name, {
            'titulo': f'Compras de {usuario.nombre_completo}',
            'usuario': usuario,
            'compras': compras,
        })


class DevolverCompraAdminView(AdministradorRequiredMixin, View):
    """
    Equivale a devolver_compra_by_id() de Administrador.php en CI3.
    """

    def get(self, request, usuario_pk, compra_pk):
        usuario = get_object_or_404(CustomUser, pk=usuario_pk)
        compra = get_object_or_404(Compra, pk=compra_pk, usuario=usuario)

        with transaction.atomic():
            nuevo_saldo = float(usuario.saldo) + float(compra.precio)

            transaccion = Transaccion.objects.create(
                usuario=usuario,
                transaccion='Reintegro',
                monto=compra.precio,
                saldo=nuevo_saldo,
            )

            from apps.comedor.models import LogCompra
            LogCompra.objects.create(
                usuario=usuario,
                dia_comprado=compra.dia_comprado,
                precio=compra.precio,
                turno=compra.turno,
                menu=compra.menu,
                tipo=compra.tipo,
                transaccion_tipo='Reintegro',
                transaccion=transaccion,
            )

            usuario.saldo = nuevo_saldo
            usuario.save(update_fields=['saldo'])
            compra.delete()

        messages.success(
            request,
            f'Se reintegró ${compra.precio} a {usuario.nombre_completo}.'
        )
        return redirect('admin_panel:ver_compras', pk=usuario_pk)


class RepartidorView(RepartidorRequiredMixin, View):
    """
    Equivale a buscar_compra_por_fecha_user() de Repartidor.php en CI3.
    """
    template_name = 'admin_panel/repartidor.html'

    def get(self, request):
        return render(request, self.template_name, {
            'titulo': 'Entrega de Viandas',
        })

    def post(self, request):
        from datetime import date
        documento = request.POST.get('numeroDni')
        hoy = date.today()

        try:
            usuario = CustomUser.objects.get(documento=int(documento))
            compra = Compra.objects.filter(
                usuario=usuario,
                dia_comprado=hoy
            ).first()
        except (CustomUser.DoesNotExist, ValueError):
            usuario = None
            compra = None

        return render(request, self.template_name, {
            'titulo': 'Entrega de Viandas',
            'usuario': usuario,
            'compra': compra,
        })


class EntregarViandaView(RepartidorRequiredMixin, View):
    """
    Equivale a entregar_compra_by_id() de Repartidor.php en CI3.
    """

    def post(self, request):
        compra_id = request.POST.get('idCompra')
        compra = get_object_or_404(Compra, pk=compra_id)

        compra.retiro = True
        compra.repartidor = request.user
        compra.save(update_fields=['retiro', 'repartidor'])

        messages.success(request, 'Vianda entregada correctamente.')
        return redirect('admin_panel:repartidor')


# apps/admin_panel/views.py - agregar al final del archivo existente


class ConfiguracionView(AdministradorRequiredMixin, View):
    """
    Equivale a configuracion_general() de Administrador.php en CI3.
    """
    template_name = 'admin_panel/configuracion.html'


    def get(self, request):
        config = Configuracion.get()
        dias_semana = {
            1: 'Lunes', 2: 'Martes', 3: 'Miércoles',
            4: 'Jueves', 5: 'Viernes', 6: 'Sábado', 7: 'Domingo'
        }
        return render(request, self.template_name, {
            'titulo': 'Configuración General',
            'config': config,
            'dias_semana': dias_semana,
        })

    def post(self, request):
        config = Configuracion.get()

        if not config:
            config = Configuracion()

        config.apertura = request.POST.get('apertura_comedor')
        config.cierre = request.POST.get('cierre_comedor')
        config.vacaciones_inicio = request.POST.get('inicio_receso')
        config.vacaciones_fin = request.POST.get('fin_receso')
        config.dia_inicial = int(request.POST.get('inicio_venta_semana'))
        config.dia_final = int(request.POST.get('fin_venta_semana'))
        config.hora_final = request.POST.get('hora_cierre_venta')
        config.permitir_ambos_turnos = request.POST.get(
            'permitir_ambos_turnos') == 'on'
        config.save()

        messages.success(request, 'Configuración guardada correctamente.')
        return redirect('admin_panel:configuracion')


class PreciosView(AdministradorRequiredMixin, View):
    """
    Equivale a configuracion_costos() de Administrador.php en CI3.
    """
    template_name = 'admin_panel/precios.html'

    def get(self, request):
        precios = Precio.objects.all()
        return render(request, self.template_name, {
            'titulo': 'Configuración de Precios',
            'precios': precios,
        })

    def post(self, request):
        precios = Precio.objects.all()
        for precio in precios:
            costo = request.POST.get(f'precio_{precio.id}')
            if costo:
                precio.costo = float(costo)
                precio.save(update_fields=['costo'])

        messages.success(request, 'Precios actualizados correctamente.')
        return redirect('admin_panel:precios')


class FeriadosView(AdministradorRequiredMixin, View):
    """
    Equivale a feriados_list() de Administrador.php en CI3.
    """
    template_name = 'admin_panel/feriados.html'

    def get(self, request):
        from datetime import date
        año = int(request.GET.get('año', date.today().year))
        feriados = Feriado.objects.filter(
            fecha__year=año
        ).order_by('fecha')

        return render(request, self.template_name, {
            'titulo': 'Feriados',
            'feriados': feriados,
            'año': año,
        })


class AgregarFeriadoView(AdministradorRequiredMixin, View):
    """
    Equivale a add_feriado() de Administrador.php en CI3.
    """

    def post(self, request):
        fecha = request.POST.get('fecha_feriado')
        detalle = request.POST.get('fecha_feriado_motivo')
        año = request.POST.get('ano')

        if fecha and detalle:
            feriado, creado = Feriado.objects.get_or_create(
                fecha=fecha,
                defaults={'detalle': detalle}
            )
            if creado:
                # Devolver compras de ese día
                self._devolver_compras_fecha(fecha, detalle)
                messages.success(request, f'Feriado {fecha} agregado.')
            else:
                messages.warning(
                    request, 'Ya existe un feriado para esa fecha.')

        return redirect(f"{request.build_absolute_uri('?')}año={año}" if año
                        else 'admin_panel:feriados')

    def _devolver_compras_fecha(self, fecha, motivo):
        """Devuelve todas las compras de una fecha dada."""
        compras = Compra.objects.filter(dia_comprado=fecha)

        for compra in compras:
            usuario = compra.usuario
            nuevo_saldo = float(usuario.saldo) + float(compra.precio)

            transaccion = Transaccion.objects.create(
                usuario=usuario,
                transaccion='Reintegro',
                monto=compra.precio,
                saldo=nuevo_saldo,
            )

            from apps.comedor.models import LogCompra
            LogCompra.objects.create(
                usuario=usuario,
                dia_comprado=compra.dia_comprado,
                precio=compra.precio,
                turno=compra.turno,
                menu=compra.menu,
                transaccion_tipo='Reintegro',
                transaccion=transaccion,
            )

            usuario.saldo = nuevo_saldo
            usuario.save(update_fields=['saldo'])
            compra.delete()

            # Email reintegro
            try:
                from django.core.mail import send_mail
                from django.template.loader import render_to_string
                from django.conf import settings

                context = {
                    'usuario': usuario,
                    'compra': compra,
                    'motivo': motivo,
                    'saldo': nuevo_saldo,
                }
                mensaje = render_to_string('emails/reintegro.html', context)
                send_mail(
                    subject=f'Reintegro por {motivo}',
                    message='',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[usuario.email],
                    html_message=mensaje,
                    fail_silently=True,
                )
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(
                    f"Error email reintegro: {e}")


class EliminarFeriadoView(AdministradorRequiredMixin, View):
    """
    Equivale a borrar_feriado() de Administrador.php en CI3.
    """

    def get(self, request, pk):
        feriado = get_object_or_404(Feriado, pk=pk)
        año = feriado.fecha.year
        feriado.delete()
        messages.success(request, 'Feriado eliminado.')
        return redirect(f"{request.build_absolute_uri('/panel/feriados/')}?año={año}")


class MenuAdminView(AdministradorRequiredMixin, View):
    """
    Equivale a updateMenu() de Vendedor.php en CI3.
    """
    template_name = 'admin_panel/menu.html'

    def get(self, request):
        menu = Menu.objects.all().order_by('dia')
        return render(request, self.template_name, {
            'titulo': 'Actualizar Menú',
            'menu': menu,
        })

    def post(self, request):
        menu = Menu.objects.all().order_by('dia')
        for item in menu:
            item.menu_basico = request.POST.get(f'basico_{item.id}', '')
            item.menu_veggie = request.POST.get(f'veggie_{item.id}', '')
            item.menu_sin_tacc = request.POST.get(f'sin_tacc_{item.id}', '')
            item.save()

        messages.success(request, 'Menú actualizado correctamente.')
        return redirect('admin_panel:menu')


class CargaCSVView(AdministradorRequiredMixin, View):
    """
    Equivale a cargar_archivo_csv() de Administrador.php en CI3.
    """
    template_name = 'admin_panel/carga_csv.html'

    def get(self, request):
        return render(request, self.template_name, {
            'titulo': 'Carga CSV',
        })

    def post(self, request):
        archivo = request.FILES.get('archivo_csv')
        separador = request.POST.get('separador', ';')

        if not archivo:
            messages.error(request, 'Seleccioná un archivo CSV.')
            return redirect('admin_panel:carga_csv')

        try:
            contenido = archivo.read().decode('utf-8')
            reader = csv.reader(io.StringIO(contenido), delimiter=separador)
            next(reader)  # saltar header
            cargas = list(reader)
        except Exception as e:
            messages.error(request, f'Error al leer el archivo: {e}')
            return redirect('admin_panel:carga_csv')

        return render(request, self.template_name, {
            'titulo': 'Carga CSV',
            'cargas': cargas,
            'separador': separador,
        })


class ConfirmarCSVView(AdministradorRequiredMixin, View):
    """
    Equivale a confirmarCargasCVS() de Administrador.php en CI3.
    """

    def post(self, request):
        errores = []
        i = 0

        while request.POST.get(f'documento_{i}'):
            documento = request.POST.get(f'documento_{i}')
            monto = float(request.POST.get(f'monto_{i}', 0))
            tipo = request.POST.get(f'tipo_{i}', 'Efectivo')

            try:
                usuario = CustomUser.objects.get(documento=int(documento))
                nuevo_saldo = float(usuario.saldo) + monto
                tipo_trans = 'Carga de Saldo' if monto >= 0 else 'Devolucion de Saldo'

                with transaction.atomic():
                    transaccion = Transaccion.objects.create(
                        usuario=usuario,
                        transaccion=tipo_trans,
                        monto=monto,
                        saldo=nuevo_saldo,
                    )
                    LogCarga.objects.create(
                        usuario=usuario,
                        vendedor=request.user,
                        monto=monto,
                        formato=tipo,
                        transaccion=transaccion,
                    )
                    usuario.saldo = nuevo_saldo
                    usuario.save(update_fields=['saldo'])

            except CustomUser.DoesNotExist:
                errores.append(documento)

            i += 1

        if errores:
            messages.warning(
                request,
                f'No se encontraron los documentos: {", ".join(errores)}'
            )
        else:
            messages.success(request, 'Cargas realizadas correctamente.')

        return redirect('admin_panel:carga_csv')


# apps/admin_panel/views.py - agregar al final


class DescargarExcelView(CajeroRequiredMixin, View):
    """
    Equivale a descargarExcel() de Vendedor.php en CI3.
    """

    def get(self, request):
        return render(request, 'admin_panel/descarga_planilla.html', {
            'titulo': 'Descargar Listados',
        })

    def post(self, request):
        fecha_str = request.POST.get('fecha')
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, 'Fecha inválida.')
            return redirect('admin_panel:excel')

        compras = Compra.objects.filter(
            dia_comprado=fecha
        ).select_related('usuario').order_by('usuario__last_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Listado'

        # Estilo header
        header_fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        header_font = Font(color='FFFFFF', bold=True)

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Listado de viandas - {fecha.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['#', 'Documento', 'Apellido',
                   'Nombre', 'Menú', 'Turno', 'Claustro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for i, compra in enumerate(compras, 1):
            ws.append([
                i,
                compra.usuario.documento,
                compra.usuario.last_name.upper(),
                compra.usuario.first_name,
                compra.get_menu_display(),
                compra.get_turno_display(),
                compra.usuario.tipo,
            ])

        # Ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Listado_{fecha}.xlsx"'
        wb.save(response)
        return response


class CierreCajaDiarioView(CajeroRequiredMixin, View):
    """
    Equivale a descargarCierreCajaDiario() de Vendedor.php en CI3.
    """

    def get(self, request):
        return render(request, 'admin_panel/descarga_informe.html', {
            'titulo': 'Informes',
        })

    def post(self, request):
        from weasyprint import HTML
        from django.template.loader import render_to_string

        fecha_str = request.POST.get('cierre_fecha')
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, 'Fecha inválida.')
            return redirect('admin_panel:informe')

        cargas = LogCarga.objects.filter(
            fecha=fecha
        ).select_related('usuario', 'vendedor', 'transaccion')

        # Totales por formato
        totales = {
            'Efectivo': {'cantidad': 0, 'total': 0},
            'Virtual': {'cantidad': 0, 'total': 0},
            'MP': {'cantidad': 0, 'total': 0},
        }

        for carga in cargas:
            formato = carga.formato
            if formato in totales:
                totales[formato]['cantidad'] += 1
                totales[formato]['total'] += float(abs(carga.monto))

        context = {
            'cargas': cargas,
            'totales': totales,
            'vendedor': request.user,
            'fecha': fecha.strftime('%d-%m-%Y'),
        }

        html_string = render_to_string(
            'admin_panel/pdf/caja_diaria.html', context)
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Cierre_{fecha}.pdf"'
        response.write(pdf)
        return response


class CierreCajaSemanalView(CajeroRequiredMixin, View):
    """
    Equivale a descargarCierreCajaSemana() de Vendedor.php en CI3.
    """

    def post(self, request):
        from weasyprint import HTML
        from django.template.loader import render_to_string

        fecha1_str = request.POST.get('cierre_fecha_1')
        fecha2_str = request.POST.get('cierre_fecha_2')

        try:
            fecha1 = datetime.strptime(fecha1_str, '%Y-%m-%d').date()
            fecha2 = datetime.strptime(fecha2_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, 'Fechas inválidas.')
            return redirect('admin_panel:informe')

        cargas = LogCarga.objects.filter(
            fecha__gte=fecha1,
            fecha__lte=fecha2,
        )

        # Armar detalle por día
        detalle = []
        fecha_actual = fecha1
        while fecha_actual <= fecha2:
            cargas_dia = cargas.filter(fecha=fecha_actual)
            dia = {
                'fecha': fecha_actual,
                'efectivo_cantidad': 0,
                'efectivo_total': 0,
                'virtual_cantidad': 0,
                'virtual_total': 0,
                'mp_cantidad': 0,
                'mp_total': 0,
            }
            for carga in cargas_dia:
                monto = float(abs(carga.monto))
                if carga.formato == 'Efectivo':
                    dia['efectivo_cantidad'] += 1
                    dia['efectivo_total'] += monto
                elif carga.formato == 'Virtual':
                    dia['virtual_cantidad'] += 1
                    dia['virtual_total'] += monto
                elif carga.formato == 'MP':
                    dia['mp_cantidad'] += 1
                    dia['mp_total'] += monto
            detalle.append(dia)
            fecha_actual += timedelta(days=1)

        context = {
            'detalle': detalle,
            'vendedor': request.user,
            'fecha1': fecha1.strftime('%d-%m-%Y'),
            'fecha2': fecha2.strftime('%d-%m-%Y'),
            'total_efectivo': sum(d['efectivo_total'] for d in detalle),
            'total_virtual': sum(d['virtual_total'] for d in detalle),
            'total_mp': sum(d['mp_total'] for d in detalle),
        }

        html_string = render_to_string(
            'admin_panel/pdf/caja_semanal.html', context)
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Cierre_{fecha1}_{fecha2}.pdf"'
        response.write(pdf)
        return response


class ResumenPedidosSemanaView(CajeroRequiredMixin, View):
    """
    Equivale a descargarResumenPedidosSemana() de Vendedor.php en CI3.
    """

    def post(self, request):
        from weasyprint import HTML
        from django.template.loader import render_to_string

        fecha1_str = request.POST.get('semana_fecha_1')
        fecha2_str = request.POST.get('semana_fecha_2')

        try:
            fecha1 = datetime.strptime(fecha1_str, '%Y-%m-%d').date()
            fecha2 = datetime.strptime(fecha2_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, 'Fechas inválidas.')
            return redirect('admin_panel:informe')

        compras = Compra.objects.filter(
            dia_comprado__gte=fecha1,
            dia_comprado__lte=fecha2,
        )

        detalle = []
        fecha_actual = fecha1
        while fecha_actual <= fecha2:
            compras_dia = compras.filter(dia_comprado=fecha_actual)
            dia = {
                'fecha': fecha_actual,
                'basico': compras_dia.filter(menu='Basico').count(),
                'veggie': compras_dia.filter(menu='Veggie').count(),
                'celiaco': compras_dia.filter(menu='Celiaco').count(),
            }
            detalle.append(dia)
            fecha_actual += timedelta(days=1)

        context = {
            'detalle': detalle,
            'fecha1': fecha1.strftime('%d-%m-%Y'),
            'fecha2': fecha2.strftime('%d-%m-%Y'),
        }

        html_string = render_to_string(
            'admin_panel/pdf/resumen_semanal.html', context)
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Pedidos_{fecha1}_{fecha2}.pdf"'
        response.write(pdf)
        return response
